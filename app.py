import streamlit as st
import io
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 1. 日期提取正则
DATE_RE = re.compile(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})')

# 2. 括号匹配正则（涵盖所有中英文/半全角括号）
BRACKET_RE = re.compile(r'（.*?）|\(.*?\)|\[.*?\]|【.*?】|\{.*?\}|［.*?］|〔.*?〕|〖.*?〗|《.*?》')

# 3. 两端剥离字符集合（包含空格、换行、各种连字符 - _ — – ~ ～ 、斜杠及常见标点）
STRIP_CHARS = ' \t\r\n\xa0\u3000-_—–~～/|\\,.，。；;：:!！?？'

# 4. 纯字符提取正则（仅保留中文、英文字母、数字）
PURE_CHAR_RE = re.compile(r'[^\u4e00-\u9fa5a-zA-Z0-9]')


def clean_source(val):
    """第一层清洗：去除括号及内容，并剥离首尾 `-` `~` 等特殊符号"""
    if not val:
        return ""
    val_str = str(val).replace('\xa0', ' ').replace('\u3000', ' ').strip()
    
    while True:
        cleaned_str = BRACKET_RE.sub('', val_str)
        cleaned_str = cleaned_str.strip(STRIP_CHARS)
        if cleaned_str == val_str:
            break
        val_str = cleaned_str
        
    return val_str


def to_pure_chars(val_str):
    """第三层兜底：提取纯文本（彻底抹去所有 - ~ _ 符号、标点与空格，转小写）"""
    if not val_str:
        return ""
    cleaned = clean_source(val_str)
    return PURE_CHAR_RE.sub('', cleaned).lower()


def match_person(source_val, name_map, norm_name_map):
    """三层递进人员匹配逻辑"""
    if not source_val:
        return None
    
    source_str = str(source_val).strip()
    
    # 【层级 1】：清洗括号与首尾 `-` `~` 符号后精确匹配
    cleaned = clean_source(source_str)
    if cleaned in name_map:
        return name_map[cleaned]
    
    # 【层级 2】：按连字符（- ~ _ / | 空格等）截取前半部分匹配（解决 CAT卡特武-华北利星行 或 CAT卡特武~930884355）
    parts = re.split(r'[-_—–/|~～\s]+', cleaned)
    if parts and parts[0]:
        first_part = parts[0].strip(STRIP_CHARS)
        if first_part in name_map:
            return name_map[first_part]
            
    # 【层级 3】：终极兜底——彻底无视 `-` `~` 标点空格匹配（解决 CAT-卡特武 vs CAT卡特武，或 ~CAT卡特武~）
    pure_source = to_pure_chars(source_str)
    if pure_source and pure_source in norm_name_map:
        return norm_name_map[pure_source]
        
    return None


def format_date_cell(cell):
    val = cell.value
    if val is None:
        return

    # 1. 已经是 datetime 对象
    if isinstance(val, datetime):
        cell.value = val.strftime("%Y/%m/%d")
        return

    # 2. 数字（Excel 内部日期序列号）
    if isinstance(val, (int, float)):
        try:
            excel_base = datetime(1899, 12, 30)
            dt = excel_base + timedelta(days=val)
            cell.value = dt.strftime("%Y/%m/%d")
        except:
            pass
        return

    # 3. 字符串处理
    if isinstance(val, str):
        val_str = val.strip()
        if not val_str:
            return

        # 优先使用正则提取年月日
        match = DATE_RE.search(val_str)
        if match:
            year, month, day = match.groups()
            try:
                dt = datetime(int(year), int(month), int(day))
                cell.value = dt.strftime("%Y/%m/%d")
                return
            except ValueError:
                pass

        # 兜底常用格式尝试
        patterns = (
            "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
            "%Y/%m/%d", "%Y-%m-%d",
            "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"
        )
        for fmt in patterns:
            try:
                dt = datetime.strptime(val_str, fmt)
                cell.value = dt.strftime("%Y/%m/%d")
                return
            except ValueError:
                continue


def process_excel(file_bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=False)
    sheet_names = wb.sheetnames

    if "留资" not in sheet_names:
        raise ValueError("工作簿中缺少“留资”工作表")

    # ===== 1. 处理“DOU+”表（仅做日期格式化，不匹配人员） =====
    if "DOU+" in sheet_names:
        dou_sheet = wb["DOU+"]
        headers_dou = [cell.value for cell in dou_sheet[1]]
        try:
            col_date = headers_dou.index("下单时间") + 1
        except ValueError:
            raise ValueError("“DOU+”表中缺少“下单时间”列")

        for row in range(2, dou_sheet.max_row + 1):
            format_date_cell(dou_sheet.cell(row, col_date))

    # ===== 2. 构建映射字典 =====
    if "名单-昵称" not in sheet_names:
        raise ValueError("工作簿中缺少“名单-昵称”工作表")

    name_sheet = wb["名单-昵称"]
    headers = [cell.value for cell in name_sheet[1]]
    try:
        col_nick = headers.index("抖音昵称") + 1
        col_emp = headers.index("员工姓名") + 1
    except ValueError:
        raise ValueError("“名单-昵称”表中缺少“抖音昵称”或“员工姓名”列")

    name_map = {}
    norm_name_map = {}  # 纯字符无符字典，专治 - ~ 杂音
    
    for row in range(2, name_sheet.max_row + 1):
        nick = name_sheet.cell(row, col_nick).value
        emp = name_sheet.cell(row, col_emp).value
        if nick and emp:
            nick_str = str(nick).strip()
            emp_str = str(emp).strip()
            
            name_map[nick_str] = emp_str
            
            cleaned_nick = clean_source(nick_str)
            if cleaned_nick:
                name_map[cleaned_nick] = emp_str
                
            pure_nick = to_pure_chars(nick_str)
            if pure_nick:
                norm_name_map[pure_nick] = emp_str

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # ===== 3. 处理“留资”表（日期格式化 + 人员映射） =====
    liu_sheet = wb["留资"]
    headers_liu = [cell.value for cell in liu_sheet[1]]

    try:
        col_date_liu = headers_liu.index("日期") + 1
        col_source_liu = headers_liu.index("客户来源") + 1
        col_person_liu = headers_liu.index("人员") + 1
    except ValueError:
        raise ValueError("“留资”表中缺少“日期”、“客户来源”或“人员”列")

    for row in range(2, liu_sheet.max_row + 1):
        # 日期格式化
        format_date_cell(liu_sheet.cell(row, col_date_liu))
        
        # 多层匹配人员
        source_val = liu_sheet.cell(row, col_source_liu).value
        person_cell = liu_sheet.cell(row, col_person_liu)
        
        matched = match_person(source_val, name_map, norm_name_map)
        
        if matched:
            person_cell.value = matched
        else:
            person_cell.fill = red_fill

    # ===== 4. 保存并返回 =====
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ----- Streamlit 界面 -----
st.set_page_config(page_title="Excel 数据转换工具（保留格式）", layout="centered")
st.title("📁 Excel 数据转换工具")
st.markdown("上传你的 Excel 文件，自动处理 DOU+（日期）、留资（日期与人员匹配），并保留所有格式。")

uploaded_file = st.file_uploader("选择 Excel 文件", type=["xlsx", "xlsm"])

if uploaded_file is not None:
    st.info(f"已上传：{uploaded_file.name} (大小：{uploaded_file.size} 字节)")
    if st.button("🚀 开始处理", type="primary"):
        with st.spinner("正在处理，请稍候..."):
            try:
                file_bytes = uploaded_file.read()
                processed_data = process_excel(file_bytes)
                st.success("✅ 处理完成！")
                st.download_button(
                    label="📥 下载处理后的 Excel 文件",
                    data=processed_data,
                    file_name=f"processed_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"❌ 处理出错：{e}")
